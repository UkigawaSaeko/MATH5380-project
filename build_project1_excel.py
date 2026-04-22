# -*- coding: utf-8 -*-
"""
从「Data for final project 1.xlsx」与 notebook 同逻辑重算回测，并输出
Project 1 要求的 Excel：多工作表 + 关键格为公式（Gross/Total/Active 相关）。

用法（在 MATH5380-project 目录下）:
  python build_project1_excel.py

输出: 同目录下 project1_results.xlsx
依赖: pandas, numpy, scipy, openpyxl
"""

from __future__ import annotations

from pathlib import Path

import numpy as np
import pandas as pd
from openpyxl import Workbook
from openpyxl.chart import LineChart, Reference
from openpyxl.styles import Font, Alignment
from openpyxl.utils import get_column_letter
from scipy.optimize import minimize


# ---------- 与 notebook 一致的函数 ----------


def nearest_psd(cov: np.ndarray, eps: float = 1e-8) -> np.ndarray:
    cov = (cov + cov.T) / 2
    vals, vecs = np.linalg.eigh(cov)
    vals = np.clip(vals, eps, None)
    return vecs @ np.diag(vals) @ vecs.T


def bl_posterior(pi, sigma, P, Q, tau=0.05, omega_scale=4.0):
    ts = tau * sigma
    mid = P @ ts @ P.T
    omega = np.diag(np.diag(mid)) * omega_scale
    inv_ts = np.linalg.inv(ts)
    inv_omega = np.linalg.inv(omega)
    m = np.linalg.inv(inv_ts + P.T @ inv_omega @ P)
    mu_bl = m @ (inv_ts @ pi + P.T @ inv_omega @ Q)
    return mu_bl, omega


def optimize_long_only(mu, sigma, delta, w_mkt, max_dev=0.05):
    n = len(mu)

    def obj(w):
        return 0.5 * delta * (w @ sigma @ w) - (mu @ w)

    cons = ({"type": "eq", "fun": lambda w: np.sum(w) - 1.0},)
    lb = np.maximum(0.0, w_mkt - max_dev)
    ub = np.minimum(1.0, w_mkt + max_dev)
    bounds = list(zip(lb, ub))
    if lb.sum() > 1.0 or ub.sum() < 1.0:
        bounds = [(0.0, 1.0)] * n
    w0 = w_mkt.copy()
    res = minimize(obj, w0, method="SLSQP", bounds=bounds, constraints=cons)
    if not res.success:
        res = minimize(obj, w0, method="SLSQP", bounds=[(0.0, 1.0)] * n, constraints=cons)
    if not res.success:
        raise RuntimeError(f"Optimization failed: {res.message}")
    w = np.clip(res.x, 0.0, 1.0)
    return w / w.sum()


def load_data_and_backtest() -> dict:
    base = Path(__file__).resolve().parent
    candidates = [
        base.parent / "Data for final project 1.xlsx",
        base / ".." / "Data for final project 1.xlsx",
        Path(r"E:\Columbia_Files\MATH5380\project\Data for final project 1.xlsx"),
    ]
    excel_path = next((p for p in candidates if p.resolve().exists()), None)
    if excel_path is None:
        raise FileNotFoundError(
            "未找到数据文件。请将「Data for final project 1.xlsx」放在项目上一级目录，或改 build_project1_excel.py 中的路径。"
        )
    excel_path = excel_path.resolve()
    print("Using data:", excel_path)

    returns_raw = pd.read_excel(excel_path, sheet_name="Index returns in USD")
    mv_raw = pd.read_excel(excel_path, sheet_name="Market values in USD")
    name_map = {
        "Bloomberg Barclays Global Inflation-Linked": "Bloomberg Barclays Global Inflation-Linked USD",
        "Bloomberg Barclays Municipal Bond": "Bloomberg Barclays US Municipal Bond",
        "Bloomberg Barclays Global High Yield": "Bloomberg Barclays Global High Yield USD",
    }
    mv_raw = mv_raw.rename(columns=name_map)

    def parse_yyyymmdd(x):
        return pd.to_datetime(str(int(x)), format="%Y%m%d")

    returns_raw["Date"] = returns_raw["Date"].apply(parse_yyyymmdd)
    mv_raw["Date"] = mv_raw["Date"].apply(parse_yyyymmdd)
    returns_raw = returns_raw.sort_values("Date").reset_index(drop=True)
    mv_raw = mv_raw.sort_values("Date").reset_index(drop=True)

    common_assets = [c for c in returns_raw.columns if c in mv_raw.columns and c != "Date"]
    returns = returns_raw[["Date"] + common_assets].copy()
    mv = mv_raw[["Date"] + common_assets].copy()
    returns = returns.set_index("Date")
    mv = mv.set_index("Date")
    assets = common_assets
    n = len(assets)
    idx = {a: i for i, a in enumerate(assets)}

    delta = 2.5
    tau = 0.05
    omega_scale = 4.0
    P = np.zeros((2, n))
    Q = np.zeros(2)
    if "Russell 1000 Growth" in idx and "MSCI World Ex USA Growth NR USD" in idx:
        P[0, idx["Russell 1000 Growth"]] = 1.0
        P[0, idx["MSCI World Ex USA Growth NR USD"]] = -1.0
        Q[0] = 0.0075
    if "Bloomberg Barclays US Aggregate" in idx and "Bloomberg Barclays Global High Yield USD" in idx:
        P[1, idx["Bloomberg Barclays US Aggregate"]] = 1.0
        P[1, idx["Bloomberg Barclays Global High Yield USD"]] = -1.0
        Q[1] = 0.004

    EXCLUDE_REBALANCE_MONTH_FROM_COV = False
    mv_year_end = mv.groupby(mv.index.year).tail(1).copy()
    available_years = sorted(mv_year_end.index.year.unique())

    weights_opt = {}
    weights_mkt = {}
    mu_pi_dict = {}
    mu_bl_dict = {}
    cov_dict = {}
    port_rets = []
    bmk_rets = []

    for y in available_years:
        reb_date = mv_year_end[mv_year_end.index.year == y].index.max()
        hold_mask = returns.index.year == (y + 1)
        if hold_mask.sum() == 0:
            continue
        hold_dates = returns.index[hold_mask]
        if EXCLUDE_REBALANCE_MONTH_FROM_COV:
            hist = returns.loc[returns.index < reb_date].tail(36)
        else:
            hist = returns.loc[returns.index <= reb_date].tail(36)
        if len(hist) < 36:
            continue
        overlap = hist.index.intersection(hold_dates)
        assert overlap.empty
        assert hold_dates.min() > reb_date
        used_for_est = hist.index.max()
        assert used_for_est <= reb_date
        assert hold_dates.min().year == y + 1

        mv_vec = mv.loc[reb_date, assets].values.astype(float)
        w_mkt = mv_vec / mv_vec.sum()
        sigma = np.cov(hist[assets].values.T, ddof=1) * 12.0
        sigma = nearest_psd(sigma)
        pi = delta * (sigma @ w_mkt)
        mu_bl, _omega = bl_posterior(pi, sigma, P, Q, tau=tau, omega_scale=omega_scale)
        w_opt = optimize_long_only(mu_bl, sigma, delta, w_mkt, max_dev=0.05)
        hold_year = y + 1
        weights_mkt[hold_year] = w_mkt
        weights_opt[hold_year] = w_opt
        mu_pi_dict[hold_year] = pi
        mu_bl_dict[hold_year] = mu_bl
        cov_dict[hold_year] = sigma

        r_next = returns.loc[hold_mask, assets]
        rp = r_next.values @ w_opt
        rb = r_next.values @ w_mkt
        port_rets.append(pd.Series(rp, index=r_next.index))
        bmk_rets.append(pd.Series(rb, index=r_next.index))

    port_rets = pd.concat(port_rets).sort_index()
    bmk_rets = pd.concat(bmk_rets).sort_index()
    final_year = max(cov_dict.keys())
    return {
        "excel_path": str(excel_path),
        "assets": assets,
        "P": P,
        "Q": Q,
        "delta": delta,
        "tau": tau,
        "omega_scale": omega_scale,
        "weights_mkt": weights_mkt,
        "weights_opt": weights_opt,
        "mu_pi_dict": mu_pi_dict,
        "mu_bl_dict": mu_bl_dict,
        "cov_dict": cov_dict,
        "port_rets": port_rets,
        "bmk_rets": bmk_rets,
        "final_year": final_year,
    }


def build_workbook(out_path: Path, d: dict) -> None:
    assets: list = d["assets"]
    P, Q = d["P"], d["Q"]
    port: pd.Series = d["port_rets"]
    bmk: pd.Series = d["bmk_rets"]
    n_months = len(port)
    last_row = 1 + n_months
    first_d, last_d = 2, last_row
    # 跨表引用必须带工作表名，否则 E/B/C/D 会指向「当前表」的错误格
    dm = "Data_Monthly!"
    rng_b = f"{dm}$B${first_d}:$B${last_row}"
    rng_c = f"{dm}$C${first_d}:$C${last_row}"
    rng_d = f"{dm}$D${first_d}:$D${last_row}"

    wb = Workbook()
    # 封面
    ws0 = wb.active
    ws0.title = "Cover"
    ws0["A1"] = "MATH 5380 — Project 1 (Long-Only Multi-Asset, Black-Litterman)"
    ws0["A1"].font = Font(bold=True, size=14)
    ws0["A3"] = "Data file:"
    ws0["B3"] = d["excel_path"]
    ws0["A4"] = "Output:"
    ws0["B4"] = str(out_path)
    ws0["A6"] = "Benchmark monthly return = same convention as code: constant year-end mkt weights within each calendar year (see report)."
    ws0["A7"] = "Statistics below use Excel formulas referencing Data_Monthly."

    # ---------- Data_Monthly ----------
    ws = wb.create_sheet("Data_Monthly", 1)
    h = [
        "Date",
        "Portfolio_Monthly_Ret",
        "Benchmark_Monthly_Ret",
        "Active_Monthly",
        "Wealth_Index_Portfolio",
        "Wealth_Index_Benchmark",
    ]
    for col, name in enumerate(h, 1):
        ws.cell(row=1, column=col, value=name)
    for i in range(n_months):
        r = 2 + i
        ts = port.index[i]
        ws.cell(row=r, column=1, value=ts)
        ws.cell(row=r, column=2, value=float(port.iloc[i]))
        ws.cell(row=r, column=3, value=float(bmk.iloc[i]))
        ws.cell(row=r, column=4, value=f"=B{r}-C{r}")
    ws.cell(row=2, column=5, value=f"=1*(1+B{first_d})")
    ws.cell(row=2, column=6, value=f"=1*(1+C{first_d})")
    for r in range(3, last_row + 1):
        ws.cell(row=r, column=5, value=f"=E{r-1}*(1+B{r})")
        ws.cell(row=r, column=6, value=f"=F{r-1}*(1+C{r})")

    def style_title(cell, s):
        cell.value = s
        cell.font = Font(bold=True, size=12)

    # ---------- Total Return Statistics ----------
    ws_t = wb.create_sheet("Total_Return_Statistics", 2)
    style_title(ws_t["A1"], "Total Return Statistics")
    ws_t["A3"] = ""
    ws_t["A4"] = "Label"
    ws_t["B4"] = "Portfolio"
    ws_t["C4"] = "Benchmark"
    ws_t["A5"] = "Geometrically annualized total return (full sample)"
    ws_t["A6"] = "Annualized volatility of total returns (monthly, ann.)"
    # 财富末值在 Data_Monthly 的 E/F 列；月数用 COUNT(B) 与财富列一致
    ws_t["B5"] = f"={dm}E{last_row}^(12/COUNT({rng_b}))-1"
    ws_t["C5"] = f"={dm}F{last_row}^(12/COUNT({rng_c}))-1"
    ws_t["B6"] = f"=STDEV.S({rng_b})*SQRT(12)"
    ws_t["C6"] = f"=STDEV.S({rng_c})*SQRT(12)"

    # ---------- Active Return Statistics ----------
    ws_a = wb.create_sheet("Active_Return_Statistics", 3)
    style_title(ws_a["A1"], "Active Return Statistics")
    ws_a["A4"] = "Arithmetically annualized active return (vs benchmark)"
    ws_a["A5"] = "Annualized volatility of active returns (tracking error)"
    ws_a["A6"] = "Information ratio"
    ws_a["B4"] = f"=AVERAGE({rng_d})*12"
    ws_a["B5"] = f"=STDEV.S({rng_d})*SQRT(12)"
    ws_a["B6"] = f"=B4/B5"

    # ---------- Gross Returns (labels + 引用 + 图) ----------
    ws_g = wb.create_sheet("Gross_Returns", 4)
    style_title(ws_g["A1"], "Gross Returns")
    ws_g["A2"] = "Growth of $1 (end-of-month wealth; values reference Data_Monthly columns E–F)"
    ws_g["A3"] = "Date"
    ws_g["B3"] = "Portfolio wealth ($1 start)"
    ws_g["C3"] = "Benchmark wealth ($1 start)"
    for r0 in range(n_months):
        r = 4 + r0
        ws_g.cell(row=r, column=1, value=f"=Data_Monthly!A{r-2}")
        ws_g.cell(row=r, column=2, value=f"=Data_Monthly!E{r-2}")
        ws_g.cell(row=r, column=3, value=f"=Data_Monthly!F{r-2}")
    g_first, g_last = 4, 3 + n_months
    if n_months:
        cht = LineChart()
        cht.title = "Gross Returns — Growth of $1"
        cht.y_axis.title = "Wealth (from $1)"
        cht.x_axis.title = "Date"
        data1 = Reference(ws_g, min_col=2, min_row=3, max_row=g_last)
        data2 = Reference(ws_g, min_col=3, min_row=3, max_row=g_last)
        cats = Reference(ws_g, min_col=1, min_row=4, max_row=g_last)
        cht.add_data(data1, titles_from_data=True)
        cht.add_data(data2, titles_from_data=True)
        cht.set_categories(cats)
        ws_g.add_chart(cht, "E2")

    # ---------- Weights_Annual ----------
    ws_w = wb.create_sheet("Weights_Annual", 5)
    style_title(ws_w["A1"], "Weights (market & optimal) — each row = hold year of portfolio")
    years = sorted(d["weights_mkt"].keys())
    for j, a in enumerate(assets, 2):
        ws_w.cell(row=2, column=j, value=a)
    ws_w["A2"] = "Year"
    ws_w["A3"] = "— Market cap (or benchmark) weights"
    for i, y in enumerate(years, 4):
        ws_w.cell(row=i, column=1, value=y)
        for j, a in enumerate(assets, 2):
            ws_w.cell(row=i, column=j, value=float(d["weights_mkt"][y][j - 2]))
    row0 = 4 + len(years) + 1
    ws_w.cell(row=row0, column=1, value="— Optimal portfolio weights")
    for i, y in enumerate(years, start=row0 + 1):
        ws_w.cell(row=i, column=1, value=y)
        for j, a in enumerate(assets, 2):
            wv = d["weights_opt"][y]
            ai = j - 2
            ws_w.cell(row=i, column=j, value=float(wv[ai]))
    n_vs = 2
    vrow = row0 + 1 + len(years) + 2
    style_title(ws_w.cell(row=vrow, column=1), "View portfolio weights (P rows) & Q (annual, relative return)")
    ws_w.cell(row=vrow + 1, column=1, value="View_ID")
    for j, a in enumerate(assets, 2):
        ws_w.cell(row=vrow + 1, column=j, value=a)
    ws_w.cell(row=vrow + 1, column=2 + len(assets), value="Q_annual")
    for k in range(n_vs):
        ws_w.cell(row=vrow + 2 + k, column=1, value=f"View {k+1}")
        for j, a in enumerate(assets, 2):
            ws_w.cell(row=vrow + 2 + k, column=j, value=float(P[k, j - 2]))
        ws_w.cell(row=vrow + 2 + k, column=2 + len(assets), value=float(Q[k]))

    # ---------- Parameters ----------
    ws_p = wb.create_sheet("Parameters", 6)
    style_title(ws_p["A1"], "Parameters (risk aversion & confidence in BL; same each year in this build)")
    ws_p["A3"] = "hold_year (portfolio applied year)"
    ws_p["B3"] = "delta (risk aversion, same in pi & MVO)"
    ws_p["C3"] = "tau (on Sigma for BL prior scale)"
    ws_p["D3"] = "omega_scale (views uncertainty multiplier; >1 = weaker views)"
    for i, y in enumerate(years, 4):
        ws_p.cell(row=i, column=1, value=y)
        ws_p.cell(row=i, column=2, value=d["delta"])
        ws_p.cell(row=i, column=3, value=d["tau"])
        ws_p.cell(row=i, column=4, value=d["omega_scale"])

    # ---------- Expected Returns ----------
    ws_e = wb.create_sheet("Expected_Returns", 7)
    style_title(ws_e["A1"], "Expected Returns — market-implied π and Black–Litterman μ_BL (annual)")

    r1 = 3
    ws_e.cell(row=r1, column=1, value="— Market-implied (equilibrium) returns π = δ Σ w_mkt")
    for j, a in enumerate(assets, 2):
        ws_e.cell(row=r1 + 1, column=j, value=a)
    ws_e.cell(row=r1 + 1, column=1, value="Year")
    for i, y in enumerate(years, r1 + 2):
        ws_e.cell(row=i, column=1, value=y)
        for j, a in enumerate(assets, 2):
            idx_a = j - 2
            ws_e.cell(row=i, column=j, value=float(d["mu_pi_dict"][y][idx_a]))
    r2 = r1 + 2 + len(years) + 2
    ws_e.cell(row=r2, column=1, value="— Black–Litterman blended E[r] μ_BL (annual)")
    for j, a in enumerate(assets, 2):
        ws_e.cell(row=r2 + 1, column=j, value=a)
    ws_e.cell(row=r2 + 1, column=1, value="Year")
    for i, y in enumerate(years, r2 + 2):
        ws_e.cell(row=i, column=1, value=y)
        for j, a in enumerate(assets, 2):
            idx_a = j - 2
            ws_e.cell(row=i, column=j, value=float(d["mu_bl_dict"][y][idx_a]))

    # ---------- Sec8 Final year ----------
    fy = d["final_year"]
    final_cov = pd.DataFrame(d["cov_dict"][fy], index=assets, columns=assets)
    final_vol = np.sqrt(np.diag(final_cov.values))
    final_cor = final_cov.values / (final_vol[:, None] * final_vol[None, :])
    w8 = wb.create_sheet("Sec08_FinalYr", 8)
    style_title(w8["A1"], f"Section 8 — Final portfolio year: {fy} (annualized)")

    w8["A3"] = "Covariances (annualized) — use same Σ as in optimization for the final year"
    w8["A3"].font = Font(bold=True, size=11)
    w8["A4"] = ""
    for j, a in enumerate(assets, 2):
        w8.cell(row=4, column=j, value=a)
    w8["A4"] = ""
    for i, a in enumerate(assets, 5):
        w8.cell(row=i, column=1, value=a)
        for j, bb in enumerate(assets, 2):
            w8.cell(row=i, column=j, value=float(final_cov.loc[a, bb]))

    wv_row = 5 + len(assets) + 1
    w8.cell(row=wv_row, column=1, value="Volatilities (annual, from final Σ)")
    w8.cell(row=wv_row, column=1).font = Font(bold=True, size=11)
    w8["A" + str(wv_row + 1)] = "Asset"
    w8["B" + str(wv_row + 1)] = "Ann. Vol (sqrt of diag in decimal variance units)"
    for k, a in enumerate(assets):
        w8.cell(row=wv_row + 2 + k, column=1, value=a)
        w8.cell(row=wv_row + 2 + k, column=2, value=float(final_vol[assets.index(a)]))

    c_row = wv_row + 2 + len(assets) + 1
    w8.cell(row=c_row, column=1, value="Correlations (from final annualized Σ)")
    w8.cell(row=c_row, column=1).font = Font(bold=True, size=11)
    for j, a in enumerate(assets, 2):
        w8.cell(row=c_row + 1, column=j, value=a)
    w8["A" + str(c_row + 1)] = ""
    for i, a in enumerate(assets, c_row + 2):
        w8.cell(row=i, column=1, value=a)
        for j, bb in enumerate(assets, 2):
            w8.cell(
                row=i, column=j, value=float(final_cor[assets.index(a), assets.index(bb)])
            )

    # ---------- Universe ----------
    wu = wb.create_sheet("Investment_Universe", 9)
    style_title(wu["A1"], "Investment universe (asset class categories)")
    for k, a in enumerate(assets, 3):
        wu.cell(row=k, column=1, value=k - 2)
        wu.cell(row=k, column=2, value=a)

    wb.save(out_path)
    print("Saved:", out_path)


def main():
    d = load_data_and_backtest()
    out = Path(__file__).resolve().parent / "project1_results.xlsx"
    build_workbook(out, d)
    print("Backtest months:", len(d["port_rets"]), "final year Σ:", d["final_year"])


if __name__ == "__main__":
    main()
